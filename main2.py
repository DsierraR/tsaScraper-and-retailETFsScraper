import matplotlib.pyplot as plt
import pandas as pd
from io import BytesIO
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.image import MIMEImage
from email.mime.application import MIMEApplication
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service as ChromeService
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook, Workbook
from datetime import datetime
import pytz
import logging
import os
import numpy as np
import time
import requests
from bs4 import BeautifulSoup
import boto3

# Set up logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Email Configuration
sender_email = "dsierraramirez115@gmail.com"
receiver_email = ["diegosierra01@yahoo.com", "arnav.ashruchi@gmail.com","jordan.valer@lmrpartners.com"]
email_password = os.environ['EMAIL_PASSWORD']

# AWS S3 Configuration
s3_client = boto3.client(
    's3',
    aws_access_key_id=os.environ['AWS_ACCESS_KEY_ID'],
    aws_secret_access_key=os.environ['AWS_SECRET_ACCESS_KEY'],
    region_name=os.environ['AWS_REGION']
)
bucket_name = 'ctabucketdata'
file_key = 'shares_outstanding_data.xlsx'

# Timezone setup
cst = pytz.timezone('America/Chicago')

# ETF Tickers
ETF_TICKERS_FIRST = ['USO', 'BNO', 'UGA']
ETF_TICKERS_SECOND = ['UCO', 'DBO', 'SCO']
BASE_URL_FIRST = "https://www.uscfinvestments.com/"
INVESTING_URLS = {
    'UCO': "https://www.investing.com/etfs/proshares-ultra-dj-ubs-crude-oil",
    'DBO': "https://www.investing.com/etfs/powershares-db-oil-fund?cid=980444",
    'SCO': "https://www.investing.com/etfs/proshares-ultrashort-dj-ubs-crude-o"
}

# Set up Selenium with simulated headful Chrome
chrome_options = Options()
chrome_options.add_argument("--headless")
chrome_options.add_argument("--disable-gpu")
chrome_options.add_argument("--window-size=1920,1080")
chrome_options.add_argument("--start-maximized")
chrome_options.add_argument("--user-agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/115.0.0.0 Safari/537.36")
chrome_options.add_argument("--disable-blink-features=AutomationControlled")
chrome_options.add_experimental_option("excludeSwitches", ["enable-automation"])
chrome_options.add_experimental_option("useAutomationExtension", False)
service = ChromeService()

# Functions to fetch shares outstanding, update the Excel file, create visualizations, and send email
def fetch_shares_outstanding_first(etf_ticker):
    url = f"{BASE_URL_FIRST}{etf_ticker.lower()}"
    logging.info(f"Fetching data for {etf_ticker}. URL: {url}")
    with webdriver.Chrome(service=service, options=chrome_options) as driver:
        driver.execute_script("Object.defineProperty(navigator, 'webdriver', {get: () => undefined})")
        driver.get(url)
        time.sleep(3)
        try:
            shares_outstanding = WebDriverWait(driver, 15).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "td[data-key='so']"))
            ).text.strip().replace(",", "")
            logging.info(f"{etf_ticker}: Shares Outstanding = {shares_outstanding}")
            return shares_outstanding
        except Exception as e:
            logging.error(f"Failed to retrieve shares outstanding for {etf_ticker}: {e}")
            return "N/A"

def fetch_shares_outstanding_static(etf_ticker):
    url = INVESTING_URLS[etf_ticker]
    headers = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"}
    response = requests.get(url, headers=headers, verify=False)
    soup = BeautifulSoup(response.content, 'html.parser')
    try:
        shares_outstanding = soup.select_one("dd[data-test='sharesOutstanding'] .key-info_dd-numeric__ZQFIs span:nth-child(2)").text.strip().replace(",", "")
        logging.info(f"{etf_ticker}: Shares Outstanding = {shares_outstanding}")
        return shares_outstanding
    except Exception as e:
        logging.error(f"Failed to retrieve shares outstanding for {etf_ticker}: {e}")
        return "N/A"

def download_excel_from_s3():
    try:
        response = s3_client.get_object(Bucket=bucket_name, Key=file_key)
        return BytesIO(response['Body'].read())
    except s3_client.exceptions.NoSuchKey:
        logging.warning(f"{file_key} not found in S3. Creating a new workbook.")
        buffer = BytesIO()
        workbook = Workbook()
        workbook.active.append(['Date'] + ETF_TICKERS_FIRST + ETF_TICKERS_SECOND)
        workbook.save(buffer)
        buffer.seek(0)
        return buffer

def upload_excel_to_s3(file_content):
    s3_client.put_object(Bucket=bucket_name, Key=file_key, Body=file_content)

def update_excel(etf_data):
    today_date = datetime.now(cst).strftime('%Y-%m-%d')
    all_tickers = ETF_TICKERS_FIRST + ETF_TICKERS_SECOND
    new_row = [today_date] + [etf_data.get(ticker, 'N/A') for ticker in all_tickers]
    excel_buffer = download_excel_from_s3()
    workbook = load_workbook(excel_buffer)
    sheet = workbook.active
    current_columns = [cell.value for cell in sheet[1]]
    missing_tickers = [ticker for ticker in all_tickers if ticker not in current_columns]
    for ticker in missing_tickers:
        sheet.cell(row=1, column=len(current_columns) + 1, value=ticker)
        current_columns.append(ticker)
    previous_row = {ticker: sheet.cell(row=sheet.max_row, column=i + 2).value for i, ticker in enumerate(all_tickers)}
    sheet.append(new_row)
    excel_buffer = BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    upload_excel_to_s3(excel_buffer)
    return new_row, previous_row

def calculate_flows(df):
    """Calculate Week over Week and Month over Month flows"""
    # Convert shares to numeric, replacing 'N/A' with NaN
    for ticker in ETF_TICKERS_FIRST + ETF_TICKERS_SECOND:
        df[ticker] = pd.to_numeric(df[ticker], errors='coerce')
    
    # Calculate total shares per day
    df['Total_Shares'] = df[ETF_TICKERS_FIRST + ETF_TICKERS_SECOND].sum(axis=1)
    
    # Calculate WoW and MoM changes
    df['WoW_Change'] = df['Total_Shares'] - df['Total_Shares'].shift(5)  # 5 business days
    df['MoM_Change'] = df['Total_Shares'] - df['Total_Shares'].shift(21)  # ~21 business days
    
    # Calculate percentage changes
    df['WoW_Pct'] = (df['WoW_Change'] / df['Total_Shares'].shift(5)) * 100
    df['MoM_Pct'] = (df['MoM_Change'] / df['Total_Shares'].shift(21)) * 100
    
    return df

def display_results(new_row, previous_row, df):
    date, *values = new_row
    print(f"\nDate: {date}")
    print("\nShares Outstanding Data:")
    print("-" * 50)
    
    for i, ticker in enumerate(ETF_TICKERS_FIRST + ETF_TICKERS_SECOND):
        prev_value = previous_row.get(ticker, 'N/A')
        current_value = values[i]
        
        if prev_value != current_value:
            change = "🔺" if float(current_value or 0) > float(prev_value or 0) else "🔻"
            print(f"{ticker}: {current_value} {change} (previous: {prev_value})")
        else:
            print(f"{ticker}: {current_value} (unchanged)")
    
    # Display flow metrics
    print("\nFlow Metrics:")
    print("-" * 50)
    latest = df.iloc[-1]
    
    if not pd.isna(latest['WoW_Change']):
        wow_arrow = "🔺" if latest['WoW_Change'] > 0 else "🔻"
        print(f"Week over Week Change: {wow_arrow} {latest['WoW_Change']:,.0f} shares ({latest['WoW_Pct']:.1f}%)")
    
    if not pd.isna(latest['MoM_Change']):
        mom_arrow = "🔺" if latest['MoM_Change'] > 0 else "🔻"
        print(f"Month over Month Change: {mom_arrow} {latest['MoM_Change']:,.0f} shares ({latest['MoM_Pct']:.1f}%)")


def create_visualization():
    excel_buffer = download_excel_from_s3()
    df = pd.read_excel(excel_buffer)
    df['Date'] = pd.to_datetime(df['Date'])
    df = calculate_flows(df)
    
    fig, ax1 = plt.subplots(figsize=(12, 7))
    
    # Left y-axis: Stacked bars
    bottom_values = np.zeros(len(df))
    colors = ['#FF9999', '#66B2FF', '#fa7832', '#FFC107', '#8BC34A', '#C09ADB']
    
    for i, ticker in enumerate(ETF_TICKERS_FIRST + ETF_TICKERS_SECOND):
        values = pd.to_numeric(df[ticker], errors='coerce').fillna(0)
        ax1.bar(df['Date'], values, bottom=bottom_values, color=colors[i], label=ticker, width=0.5)
        bottom_values += values
    
    ax1.plot(df['Date'], df['Total_Shares'], label='Total Shares', color='black', linewidth=2)
    ax1.set_xlabel('Date')
    ax1.set_ylabel('Shares Outstanding', color='black')
    
    # Right y-axis: Flow changes
    ax2 = ax1.twinx()
    ax2.plot(df['Date'], df['WoW_Change'], label='WoW Flow', color='blue', linestyle='--')
    ax2.plot(df['Date'], df['MoM_Change'], label='MoM Flow', color='red', linestyle='--')
    ax2.axhline(y=0, color='gray', linestyle='-', alpha=0.3)
    ax2.set_ylabel('Change in Shares', color='gray')
    
    lines1, labels1 = ax1.get_legend_handles_labels()
    lines2, labels2 = ax2.get_legend_handles_labels()
    ax1.legend(lines1 + lines2, labels1 + labels2, loc='upper left', fontsize='small')
    
    plt.title('ETF Shares Outstanding and Flows Over Time')
    fig.autofmt_xdate(rotation=45)
    plt.tight_layout()
    
    img_buffer = BytesIO()
    plt.savefig(img_buffer, format='png', bbox_inches='tight')
    img_buffer.seek(0)
    plt.close(fig)
    
    return img_buffer, df

def send_email_with_visualization(new_row, previous_row):
    date, *values = new_row
    message = MIMEMultipart("related")
    message["From"] = sender_email
    message["To"] = ", ".join(receiver_email)
    message["Subject"] = "Retail Oil ETFs - Shares Outstanding Data"
    
    img_buffer, df = create_visualization()
    latest = df.iloc[-1]
    
    body = f"""<p>Date: {date}</p>
    <h3>Daily Changes:</h3>"""
    
    changes = False
    for i, ticker in enumerate(ETF_TICKERS_FIRST + ETF_TICKERS_SECOND):
        prev_value = previous_row.get(ticker)
        curr_value = values[i]
        if prev_value != curr_value:
            arrow = "🔺" if float(curr_value or 0) > float(prev_value or 0) else "🔻"
            body += f"<p>{ticker}: {curr_value} {arrow} (previous: {prev_value})</p>"
            changes = True
    
    if not changes:
        body += "<p>No changes in shares outstanding today.</p>"
    
    body += "<h3>Flow Metrics:</h3>"
    if not pd.isna(latest['WoW_Change']):
        arrow = "🔺" if latest['WoW_Change'] > 0 else "🔻"
        body += f"<p>Week over Week Change: {arrow} {latest['WoW_Change']:,.0f} shares ({latest['WoW_Pct']:.1f}%)</p>"
    if not pd.isna(latest['MoM_Change']):
        arrow = "🔺" if latest['MoM_Change'] > 0 else "🔻"
        body += f"<p>Month over Month Change: {arrow} {latest['MoM_Change']:,.0f} shares ({latest['MoM_Pct']:.1f}%)</p>"
    
    body += "<h3>Visualization:</h3><br><img src='cid:visualization'><br>"
    
    message.attach(MIMEText(body, "html"))
    
    img = MIMEImage(img_buffer.getvalue())
    img.add_header("Content-ID", "<visualization>")
    message.attach(img)
    
    excel_buffer = download_excel_from_s3()
    excel_attachment = MIMEApplication(excel_buffer.getvalue(), _subtype="xlsx")
    excel_attachment.add_header("Content-Disposition", "attachment", filename="shares_outstanding_data.xlsx")
    message.attach(excel_attachment)
    
    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(sender_email, email_password)
            server.send_message(message)
        logging.info("Email with visualization and Excel attachment sent successfully.")
    except Exception as e:
        logging.error(f"An error occurred while sending the email: {e}")

def main():
    etf_data = {ticker: fetch_shares_outstanding_first(ticker) for ticker in ETF_TICKERS_FIRST}
    etf_data.update({ticker: fetch_shares_outstanding_static(ticker) for ticker in ETF_TICKERS_SECOND})
    new_row, previous_row = update_excel(etf_data)
    send_email_with_visualization(new_row, previous_row)

if __name__ == "__main__":
    main()
